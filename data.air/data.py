# -*- encoding=utf8 -*-
__author__ = "Administrator"

from airtest.core.api import *

auto_setup(__file__)

import openpyxl
'''
#创建表格，写入数据
wd=openpyxl.Workbook()#操作表格函数，创建一个表格，注意大写
sheet=wd.active#当前活动页签
sheet.title='手机麻麻参数'#将当前页签改名
wd.create_sheet()#创建表
value=[['用例ID','TID','用例描述','元素动作','元素识别','元素名称','事件动作数据','休息'],
['1','swipe','查找并点击app','划动屏幕','屏幕比例','左边的位置','0.1','0'],
['2','swipe','查找并点击app','划动屏幕','屏幕比例','右边的位置','0.9','0'],
['3','login','线上环境登录','登录账号','name','com.jike.shoujimama:id/phoneET','13422883445','0'],
['4','login','线上环境登录','登录密码','name','com.jike.shoujimama:id/et_setting_psw','123456','0']]
for i in range(len(value)):#value值里的每个表遍历一遍
    for j in range(len(value[i])):#把表里的每个表又遍历一遍
        sheet.cell(row=i+1,column=j+1,value=str(value[i][j]))#写入数据
wd.save("手机麻麻自动化文档.xlsx")#保存文件

#打开表格，读取值
wb=openpyxl.load_workbook('手机麻麻.xlsx')#打开表
sheet=wb.get_sheet_by_name('手机麻麻参数')#读取当前表名为手机麻麻参数
for i in sheet.rows:#遍历当前页签行的所有行
    for j in i :#遍历每一行的值
        print (j.luvae,'\b',end='')#打印值，end=''表示不换行
    print()#print sheet.row的语句，显示每一行
 
#打开表格，读取某一行的值
wb=openpyxl.load_workbook('D:\AirTestsjmama\手机麻麻自动化文档.xlsx')#打开表
sheet=wb['手机麻麻参数']#读取当前表名为手机麻麻参数,get_sheet_by_name()方法已经弃用
row=sheet.rows

for i in row:
    if(i[0].value != 1):#判断第一个是不是value,不是不继续（只输出是的）
        continue
    for j in i :
        print(j.value,'\b',end='')#再判断之后的语句输出该行，print结果是该行的所值
'''

wb=openpyxl.load_workbook('D:\AirTestsjmama\手机麻麻自动化文档.xlsx')#打开表
sheet=wb['手机麻麻参数']#读取当前表名为手机麻麻参数,get_sheet_by_name()方法已经弃用
row=sheet.rows
def rightcoords():
    for i in row:
        if(i[5].value == '左边的位置'):#判断i(每一行)第5列的位置
            print(i[6].value)#然后输出这一行的第7列的值

def leftcoords():
    for i in row:
        if(i[5].value == '右边的位置'):#判断i(每一行)第5列的位置
            print(i[6].value)#然后输出这一行的第7列的值

rightcoords()
leftcoords()